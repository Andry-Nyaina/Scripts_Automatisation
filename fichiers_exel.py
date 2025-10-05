import openpyxl

wb1 = openpyxl.load_workbook("octobre.xlsx", data_only=True)
wb2 = openpyxl.load_workbook("novembre.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("decembre.xlsx", data_only=True)

def add_data_in_wb(wb, d):
    sheet = wb.active

    for row in range(2, sheet.max_row+1):
        nom_article = sheet.cell(row, 1).value
        if not nom_article:
            break
        total_vente = sheet.cell(row, 4).value
        if d.get(nom_article):
            d[nom_article].append(total_vente)
        else:
            d[nom_article] = [total_vente]

donnees = {}
add_data_in_wb(wb1, donnees)
add_data_in_wb(wb2, donnees)
add_data_in_wb(wb3, donnees)
print(donnees)

